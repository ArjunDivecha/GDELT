#!/usr/bin/env python3
from __future__ import annotations

import argparse
from copy import copy
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook


COUNTRY_BUCKETS = [
    ("Singapore", "SGP"),
    ("Australia", "AUS"),
    ("Canada", "CAN"),
    ("Germany", "DEU"),
    ("Japan", "JPN"),
    ("Switzerland", "CHE"),
    ("U.K.", "GBR"),
    ("U.S. NASDAQ", "USA"),
    ("U.S.", "USA"),
    ("France", "FRA"),
    ("Netherlands", "NLD"),
    ("Sweden", "SWE"),
    ("Italy", "ITA"),
    ("China A", "CHN"),
    ("Chile", "CHL"),
    ("Indonesia", "IDN"),
    ("Philippines", "PHL"),
    ("Poland", "POL"),
    ("US SmallCap", "USA"),
    ("Malaysia", "MYS"),
    ("Taiwan", "TWN"),
    ("Mexico", "MEX"),
    ("Korea", "KOR"),
    ("Brazil", "BRA"),
    ("South Africa", "ZAF"),
    ("Denmark", "DNK"),
    ("India", "IND"),
    ("China H", "CHN"),
    ("Hong Kong", "HKG"),
    ("Thailand", "THA"),
    ("Turkey", "TUR"),
    ("Spain", "ESP"),
    ("Vietnam", "VNM"),
    ("Saudi Arabia", "SAU"),
]

INDICATORS = [
    ("country_news_sentiment", "country_news_sentiment"),
    ("country_news_risk", "country_news_risk"),
    ("country_news_sentiment_raw", "country_news_sentiment_raw"),
    ("country_news_risk_raw", "country_news_risk_raw"),
    ("country_news_attention", "country_news_attention"),
    ("local_attention_share", "local_attention_share"),
    ("sentiment_x_attention", "country_news_sentiment_x_attention"),
    ("local_tone", "local_tone"),
    ("foreign_tone", "foreign_tone"),
    ("attention_shock", "attention_shock"),
    ("tone_dispersion", "tone_dispersion"),
    ("tone_wavg_wordcount", "tone_wavg_wordcount"),
    ("tone_mean", "tone_mean"),
    ("tone_p50", "tone_p50"),
    ("positive_mean", "positive_mean"),
    ("negative_mean", "negative_mean"),
    ("polarity_mean", "polarity_mean"),
    ("n_articles", "n_articles"),
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export GDELT country-day aggregates into a workbook in the sample country layout."
    )
    parser.add_argument(
        "--aggregates-dir",
        default="data/aggregates",
        help="Directory containing daily aggregate folders with country_day_all.csv files.",
    )
    parser.add_argument(
        "--panel-csv",
        default="",
        help="Optional prebuilt panel CSV. If provided, this is used instead of data/aggregates/*/country_day_all.csv.",
    )
    parser.add_argument(
        "--panel-parquet",
        default="",
        help="Optional prebuilt panel parquet. Preferred for the stream-first pipeline.",
    )
    parser.add_argument(
        "--output",
        default="output/spreadsheet/gdelt_country_sentiment.xlsx",
        help="Path to the output workbook.",
    )
    parser.add_argument(
        "--template-xlsx",
        default="",
        help="Optional workbook to borrow styles from.",
    )
    return parser.parse_args()


def load_panel(aggregates_dir: Path, panel_csv: str, panel_parquet: str) -> pd.DataFrame:
    if panel_parquet:
        frame = pd.read_parquet(panel_parquet)
    elif panel_csv:
        frame = pd.read_csv(panel_csv)
    else:
        parquet_files = sorted(aggregates_dir.glob("*.parquet"))
        if parquet_files:
            frame = pd.concat((pd.read_parquet(path) for path in parquet_files), ignore_index=True)
        else:
            files = sorted(aggregates_dir.glob("*/country_day_all.csv"))
            if not files:
                raise FileNotFoundError(
                    f"No country_day_all.csv files or parquet panel files found under {aggregates_dir}"
                )
            frame = pd.concat((pd.read_csv(path) for path in files), ignore_index=True)
    frame = frame.dropna(subset=["country_iso3"]).copy()
    frame["date"] = pd.to_datetime(frame["date"]).dt.normalize()
    frame = frame.sort_values(["date", "country_iso3"]).drop_duplicates(
        subset=["date", "country_iso3"], keep="last"
    )
    return frame


def make_style_book(template_path: str) -> tuple[Workbook, dict[str, object]]:
    if template_path:
        template_wb = load_workbook(template_path)
        template_ws = template_wb[template_wb.sheetnames[0]]
        style_refs = {
            "header_style": copy(template_ws["B1"]._style),
            "date_style": copy(template_ws["A2"]._style),
            "value_style": copy(template_ws["B2"]._style),
            "header_font": copy(template_ws["B1"].font),
            "date_font": copy(template_ws["A2"].font),
            "value_font": copy(template_ws["B2"].font),
            "header_fill": copy(template_ws["B1"].fill),
            "date_fill": copy(template_ws["A2"].fill),
            "value_fill": copy(template_ws["B2"].fill),
            "header_border": copy(template_ws["B1"].border),
            "date_border": copy(template_ws["A2"].border),
            "value_border": copy(template_ws["B2"].border),
            "header_alignment": copy(template_ws["B1"].alignment),
            "date_alignment": copy(template_ws["A2"].alignment),
            "value_alignment": copy(template_ws["B2"].alignment),
            "header_number_format": template_ws["B1"].number_format,
            "date_number_format": template_ws["A2"].number_format,
            "value_number_format": template_ws["B2"].number_format,
            "column_a_width": template_ws.column_dimensions["A"].width,
        }
        wb = Workbook()
        wb.remove(wb.active)
        return wb, style_refs

    wb = Workbook()
    wb.remove(wb.active)
    style_refs = {
        "column_a_width": 10.33,
        "date_number_format": "mm-dd-yy",
        "value_number_format": "0.000000",
    }
    return wb, style_refs


def apply_template_style(cell, prefix: str, style_refs: dict[str, object]) -> None:
    style_key = f"{prefix}_style"
    if style_key in style_refs:
        cell._style = copy(style_refs[style_key])
        cell.font = copy(style_refs[f"{prefix}_font"])
        cell.fill = copy(style_refs[f"{prefix}_fill"])
        cell.border = copy(style_refs[f"{prefix}_border"])
        cell.alignment = copy(style_refs[f"{prefix}_alignment"])
        cell.number_format = style_refs[f"{prefix}_number_format"]


def populate_sheet(
    ws, wide: pd.DataFrame, sheet_name: str, indicator: str, style_refs: dict[str, object]
) -> None:
    ws.title = sheet_name
    ws.column_dimensions["A"].width = style_refs["column_a_width"]

    ws.cell(row=1, column=1, value=None)
    for idx, (label, _iso3) in enumerate(COUNTRY_BUCKETS, start=2):
        cell = ws.cell(row=1, column=idx, value=label)
        apply_template_style(cell, "header", style_refs)

    for row_idx, dt in enumerate(wide.index, start=2):
        date_cell = ws.cell(row=row_idx, column=1, value=dt.to_pydatetime())
        apply_template_style(date_cell, "date", style_refs)
        if "date_number_format" in style_refs:
            date_cell.number_format = style_refs["date_number_format"]

        for col_idx, (label, iso3) in enumerate(COUNTRY_BUCKETS, start=2):
            value = wide.at[dt, iso3] if iso3 in wide.columns else None
            if pd.isna(value):
                value = None
            value_cell = ws.cell(row=row_idx, column=col_idx, value=value)
            apply_template_style(value_cell, "value", style_refs)
            if "value_number_format" in style_refs and indicator != "n_articles":
                value_cell.number_format = style_refs["value_number_format"]
            if indicator == "n_articles":
                value_cell.number_format = "0"


def build_indicator_panel(frame: pd.DataFrame, indicator: str) -> pd.DataFrame:
    panel = frame.pivot(index="date", columns="country_iso3", values=indicator)
    panel = panel.sort_index()
    needed = sorted({iso3 for _label, iso3 in COUNTRY_BUCKETS})
    panel = panel.reindex(columns=needed)
    return panel


def add_readme_sheet(wb: Workbook, frame: pd.DataFrame) -> None:
    ws = wb.create_sheet("README", 0)
    min_date = frame["date"].min().date().isoformat()
    max_date = frame["date"].max().date().isoformat()
    available_indicators = [sheet_name for sheet_name, indicator in INDICATORS if indicator in frame.columns]
    lines = [
        "GDELT country sentiment workbook",
        f"Coverage in workbook: {min_date} to {max_date}",
        "Layout: dates down column A, country buckets across row 1.",
        "Aliases: U.S., U.S. NASDAQ, and US SmallCap all map to USA.",
        "Aliases: China A and China H both map to CHN.",
        "Source data: panel CSV or data/aggregates/*/country_day_all.csv",
        "Workbook sheets:",
    ]
    for idx, line in enumerate(lines, start=1):
        ws.cell(row=idx, column=1, value=line)
    start_row = len(lines) + 1
    for offset, sheet_name in enumerate(available_indicators, start=start_row):
        ws.cell(row=offset, column=1, value=sheet_name)


def main() -> None:
    args = parse_args()
    aggregates_dir = Path(args.aggregates_dir)
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    frame = load_panel(aggregates_dir, args.panel_csv, args.panel_parquet)
    wb, style_refs = make_style_book(args.template_xlsx)
    add_readme_sheet(wb, frame)

    for sheet_name, indicator in INDICATORS:
        if indicator not in frame.columns:
            continue
        wide = build_indicator_panel(frame, indicator)
        ws = wb.create_sheet(sheet_name)
        populate_sheet(ws, wide, sheet_name, indicator, style_refs)

    wb.save(output_path)
    print(f"saved {output_path}")
    print(f"rows={frame['date'].nunique()} dates, countries={len(COUNTRY_BUCKETS)} buckets")


if __name__ == "__main__":
    main()
