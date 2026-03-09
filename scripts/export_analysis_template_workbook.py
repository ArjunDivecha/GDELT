#!/usr/bin/env python3
from __future__ import annotations

import argparse
from copy import copy
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell


VARIABLE_SHEETS = [
    ("px_last", "px_last"),
    ("ret_1d", "ret_1d"),
    ("ret_fwd_1d", "ret_fwd_1d"),
    ("ret_fwd_5d", "ret_fwd_5d"),
    ("ret_fwd_20d", "ret_fwd_20d"),
    ("ret_fwd_1session", "ret_fwd_1session"),
    ("ret_fwd_5session", "ret_fwd_5session"),
    ("ret_fwd_20session", "ret_fwd_20session"),
    ("n_articles", "n_articles"),
    ("local_n_articles", "local_n_articles"),
    ("foreign_n_articles", "foreign_n_articles"),
    ("unknown_source_n_articles", "unknown_source_n_articles"),
    ("local_source_total_articles", "local_source_total_articles"),
    ("source_resolution_rate", "source_resolution_rate"),
    ("local_attention_share", "local_attention_share"),
    ("tone_mean", "tone_mean"),
    ("tone_wavg_wordcount", "tone_wavg_wordcount"),
    ("negative_mean", "negative_mean"),
    ("tone_dispersion", "tone_dispersion"),
    ("local_tone", "local_tone"),
    ("foreign_tone", "foreign_tone"),
    ("country_news_sentiment_raw", "country_news_sentiment_raw"),
    ("country_news_attention", "country_news_attention"),
    ("sentiment_x_attention_raw", "sentiment_x_attention_raw"),
    ("country_news_risk_raw", "country_news_risk_raw"),
    ("attention_shock", "attention_shock"),
    ("local_tone_z", "local_tone_z"),
    ("foreign_tone_z", "foreign_tone_z"),
    ("tone_dispersion_z", "tone_dispersion_z"),
    ("local_attention_share_z", "local_attention_share_z"),
    ("country_news_sentiment", "country_news_sentiment"),
    ("country_news_sent_x_attention", "country_news_sentiment_x_attention"),
    ("country_news_risk", "country_news_risk"),
    ("days_since_prior_observation", "days_since_prior_observation"),
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Export the merged signal/return panel into an Excel workbook that matches the "
            "Daily Return.xlsx row and column layout exactly."
        )
    )
    parser.add_argument(
        "--template-xlsx",
        default="Daily Return.xlsx",
        help="Template workbook whose first sheet defines the exact date rows and bucket columns.",
    )
    parser.add_argument(
        "--backtest-panel-parquet",
        default="data/panels/country_signal_backtest_daily.parquet",
        help="Merged signal/return parquet panel with date and bucket_label columns.",
    )
    parser.add_argument(
        "--output-xlsx",
        default="output/spreadsheet/gdelt_analysis_template_workbook.xlsx",
        help="Destination workbook path.",
    )
    parser.add_argument(
        "--variables",
        default="",
        help=(
            "Optional comma-separated subset of variables or sheet names to export. "
            "If omitted, all configured variables are exported."
        ),
    )
    return parser.parse_args()


def load_template_grid(path: Path) -> tuple[list[str], pd.DatetimeIndex, dict[str, object]]:
    workbook = load_workbook(path)
    sheet = workbook[workbook.sheetnames[0]]

    headers = [sheet.cell(1, column).value for column in range(1, sheet.max_column + 1)]
    if not headers or headers[0] != "Country":
        raise ValueError("Template workbook first sheet must have 'Country' in cell A1")

    bucket_labels = headers[1:]
    dates = []
    for row in range(2, sheet.max_row + 1):
        value = sheet.cell(row, 1).value
        if value is None:
            raise ValueError(f"Template workbook has a blank date cell at row {row}")
        dates.append(pd.Timestamp(value).normalize())

    column_widths = {}
    for column in range(1, sheet.max_column + 1):
        letter = sheet.cell(1, column).column_letter
        width = sheet.column_dimensions[letter].width
        if width is not None:
            column_widths[letter] = width

    styles = {
        "header": copy(sheet["A1"]._style),
        "date": copy(sheet["A2"]._style),
        "value": copy(sheet["B2"]._style),
        "column_widths": column_widths,
    }

    return bucket_labels, pd.DatetimeIndex(dates, name="date"), styles


def load_backtest_panel(path: Path) -> pd.DataFrame:
    frame = pd.read_parquet(path)
    required = {"date", "bucket_label"}
    missing = sorted(required - set(frame.columns))
    if missing:
        raise ValueError(f"Backtest panel missing required columns: {', '.join(missing)}")

    frame = frame.copy()
    frame["date"] = pd.to_datetime(frame["date"]).dt.normalize()
    frame = frame.sort_values(["date", "bucket_label"]).drop_duplicates(
        subset=["date", "bucket_label"], keep="last"
    )
    return frame


def parse_variable_list(value: str) -> list[str]:
    return [part.strip() for part in value.split(",") if part.strip()]


def build_wide_panel(
    frame: pd.DataFrame,
    variable: str,
    template_dates: pd.DatetimeIndex,
    bucket_labels: list[str],
) -> pd.DataFrame:
    wide = frame.pivot(index="date", columns="bucket_label", values=variable)
    wide = wide.reindex(index=template_dates, columns=bucket_labels)
    return wide


def styled_cell(sheet, value, style) -> WriteOnlyCell:
    cell = WriteOnlyCell(sheet, value=value)
    cell._style = copy(style)
    return cell


def write_variable_sheet(
    workbook: Workbook,
    sheet_name: str,
    wide: pd.DataFrame,
    bucket_labels: list[str],
    styles: dict[str, object],
) -> None:
    sheet = workbook.create_sheet(title=sheet_name)
    for letter, width in styles["column_widths"].items():
        sheet.column_dimensions[letter].width = width

    header_row = [styled_cell(sheet, "Country", styles["header"])]
    for label in bucket_labels:
        header_row.append(styled_cell(sheet, label, styles["header"]))
    sheet.append(header_row)

    for date_value, values in wide.iterrows():
        row = [styled_cell(sheet, date_value.to_pydatetime(), styles["date"])]
        for value in values.tolist():
            if pd.isna(value):
                row.append(None)
            else:
                scalar = value.item() if hasattr(value, "item") else value
                row.append(styled_cell(sheet, scalar, styles["value"]))
        sheet.append(row)


def main() -> None:
    args = parse_args()
    template_path = Path(args.template_xlsx)
    panel_path = Path(args.backtest_panel_parquet)
    output_path = Path(args.output_xlsx)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    bucket_labels, template_dates, styles = load_template_grid(template_path)
    frame = load_backtest_panel(panel_path)
    requested_variables = parse_variable_list(args.variables)

    configured_variables = [
        (sheet_name, variable)
        for sheet_name, variable in VARIABLE_SHEETS
        if variable in frame.columns
    ]
    if not configured_variables:
        raise ValueError("No configured variables found in the backtest panel.")

    if requested_variables:
        requested_set = set(requested_variables)
        available_variables = [
            (sheet_name, variable)
            for sheet_name, variable in configured_variables
            if sheet_name in requested_set or variable in requested_set
        ]
        matched_keys = {sheet_name for sheet_name, _variable in available_variables} | {
            variable for _sheet_name, variable in available_variables
        }
        missing = [value for value in requested_variables if value not in matched_keys]
        if missing:
            raise ValueError(
                "Requested variables not available for export: " + ", ".join(missing)
            )
    else:
        available_variables = configured_variables

    workbook = Workbook(write_only=True)

    for sheet_name, variable in available_variables:
        wide = build_wide_panel(frame, variable, template_dates=template_dates, bucket_labels=bucket_labels)
        write_variable_sheet(
            workbook,
            sheet_name=sheet_name,
            wide=wide,
            bucket_labels=bucket_labels,
            styles=styles,
        )

    workbook.save(output_path)
    print(f"saved {output_path}")
    print(
        f"sheets={len(available_variables)} rows_per_sheet={len(template_dates) + 1} "
        f"bucket_columns={len(bucket_labels)}"
    )


if __name__ == "__main__":
    main()
